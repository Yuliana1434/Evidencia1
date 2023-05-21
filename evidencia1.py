from prettytable import PrettyTable
import os
import sys
import csv
import openpyxl 
import datetime
import sqlite3
from sqlite3 import Error

if (os.path.exists("Libreria.db")):
    print ("Base de datos cargada")
else:
    # Crear una conexión a la base de datos
    try:
        conn = sqlite3.connect("Libreria.db")
        cursor = conn.cursor()
        # Crear la tabla "autores"
        cursor.execute("CREATE TABLE IF NOT EXISTS autores (id_autor INTEGER PRIMARY KEY, nombre_autor TEXT);")
        # Crear la tabla "generos"
        cursor.execute("CREATE TABLE IF NOT EXISTS generos (id_genero INTEGER PRIMARY KEY, nombre_genero TEXT);")
        # Crear la tabla "libros" con las relaciones
        cursor.execute("CREATE TABLE IF NOT EXISTS libros (id_libro INTEGER PRIMARY KEY,titulo TEXT,autor_id INTEGER,genero_id INTEGER,año_publicacion TIMESTAMP,isbn INTEGER(10),fecha_adquisicion TIMESTAMP,FOREIGN KEY (autor_id) REFERENCES autores (id_autor),FOREIGN KEY (genero_id) REFERENCES generos (id_genero));")
        print("Tablas creadas exitosamente")
    except Error as e:
        print(e)
    finally:
        # Cerrar la conexión a la base de datos
        conn.close()


while True:
    print("")
    print("Menu principal")
    print("")
    print("1.-Registrar Nuevo Ejemplar\n2.-Consulta y Reportes\n3.-Agregar Autor\n4.-Agregar genero\n5.-Salir")
    opcion=int(input("¿Que opcion desea?: "))
        
    if (opcion==1):
        try:
            with sqlite3.connect("Libreria.db") as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM autores LIMIT 1;")
                autor_row = cursor.fetchone()  # Obtiene el primer registro de la tabla

                cursor.execute("SELECT * FROM generos LIMIT 1;")
                genero_row = cursor.fetchone()  # Obtiene el primer registro de la tabla
                if autor_row is None and genero_row is None:
                    print()
                    print("No hay datos en las tablas de autores y géneros.")
                else:
                    # Capturar datos de la tabla libros
                    while True:
                        print("")
                        print("Captura de Datos de Libros")
                        print("")
                        # Capturar el título del libro
                        titulo = input("Ingrese el título del libro: ")
                        if titulo.strip() == "":
                            continue

                        # Capturar el ID del autor
                        while True:
                            cursor.execute("SELECT * FROM autores;")
                            autor_rows = cursor.fetchall()  # Obtiene el registro 
                            tabla_autores = PrettyTable()
                            # Definir las columnas de la tabla
                            tabla_autores.field_names = ["ID", "Nombre"]
                            # Agregar el registro de autor_row a la tabla
                            for autor_row in autor_rows:
                                tabla_autores.add_row(autor_row)
                            # Imprimir la tabla
                            print(tabla_autores)

                            autor_id = input("Ingrese el ID del autor: ")
                            cursor.execute("SELECT * FROM autores WHERE id_autor = ?;", (autor_id,))
                            autor_row = cursor.fetchone()
                            if autor_row is None:
                                print("El ID del autor no existe. Por favor, ingrese un ID válido.")
                                continue
                            else:
                                break

                        # Capturar el ID del género
                        while True:
                            cursor.execute("SELECT * FROM generos;")
                            genero_rows = cursor.fetchall()  # Obtiene todos los registros

                            # Crear una instancia de PrettyTable para los géneros
                            tabla_generos = PrettyTable()
                            tabla_generos.field_names = ["ID", "Nombre"]

                            # Agregar los registros de género a la tabla
                            for genero_row in genero_rows:
                                tabla_generos.add_row(genero_row)

                            # Imprimir la tabla de géneros
                            print(tabla_generos)

                            genero_id = input("Ingrese el ID del género: ")
                            cursor.execute("SELECT * FROM generos WHERE id_genero = ?;", (genero_id,))
                            genero_row = cursor.fetchone()  # Obtiene el registro con el ID proporcionado

                            if genero_row is None:
                                print("El ID del género no existe. Por favor, ingrese un ID válido.")
                                continue
                            else:
                                break


                        # Capturar el año de publicación
                        while True:
                            año_publicacion_str = input("Ingrese la fecha de publicación en formato dd/mm/yyyy: ")
                            try:
                                año_publicacion = datetime.datetime.strptime(año_publicacion_str, "%d/%m/%Y")
                                break
                            except ValueError:
                                print("Formato de fecha incorrecto. Ingrese la fecha en el formato especificado.")
                                continue

                        # Capturar el ISBN
                        while True:
                            isbn = input("Ingrese el ISBN (10 dígitos): ")
                            # Por ejemplo, puedes verificar si el ISBN tiene la longitud correcta
                            if len(isbn) != 10:
                                print("El ISBN debe tener exactamente 10 dígitos.")
                                continue
                            else:
                                break

                        # Capturar la fecha de adquisición
                        while True:
                            fecha_adquisicion_str = input("Ingrese la fecha de adquisición en formato dd/mm/yyyy: ")
                            try:
                                fecha_adquisicion = datetime.datetime.strptime(fecha_adquisicion_str, "%d/%m/%Y")
                                break
                            except ValueError:
                                print("Formato de fecha incorrecto. Ingrese la fecha en el formato especificado.")
                        
                        # Realizar la inserción en la tabla libros
                        try:
                            cursor.execute("INSERT INTO libros (titulo, autor_id, genero_id, año_publicacion, isbn, fecha_adquisicion) VALUES (?, ?, ?, ?, ?, ?);",
                                        (titulo, autor_id, genero_id, año_publicacion, isbn, fecha_adquisicion))
                            conn.commit()
                            print("El libro se ha registrado exitosamente.")
                            break
                        except Error as e:
                            print("Error al insertar el libro:", e)
                            break
                        except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        except Error as e:
            print (e)
        except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        finally:
            conn.close()

    if (opcion==2):
        while True:
            print("")
            print("Menu Consultas y Reportes")
            print("")
            print("1.-Consulta del titulo\n2.-Reportes\n3.-Menu principal")
            opcion2=int(input("Seleccione una opcion: "))
                
            if (opcion2==1):
                while True:
                    print("")
                    print("Menu Consulta Titulo")
                    print("1.-Por Titulo\n2.-Por ISBN\n3.-Menu Anterior")
                    opcion3=int(input("Seleccione una opcion: "))
                    if opcion3 == 1:
                        try:
                            with sqlite3.connect("Libreria.db") as conn:
                                cursor = conn.cursor()
                                titulo_buscar = input("Ingrese el título a buscar: ")

                                # Realizar la consulta para obtener los libros con el título especificado y sus detalles con el JOIN
                                cursor.execute("SELECT libros.titulo, autores.nombre_autor, generos.nombre_genero, strftime('%d/%m/%Y', libros.año_publicacion), libros.isbn, strftime('%d/%m/%Y', libros.fecha_adquisicion) FROM libros JOIN autores ON libros.autor_id = autores.id_autor JOIN generos ON libros.genero_id = generos.id_genero WHERE libros.titulo LIKE ?;", ('%' + titulo_buscar + '%',))
                                libros_encontrados = cursor.fetchall()

                                if libros_encontrados:
                                    tabla_libros = PrettyTable(["Título", "Autor", "Género", "Año de Publicación", "ISBN", "Fecha de Adquisición"])

                                    # Agregar los registros de libros encontrados a la tabla
                                    for libro in libros_encontrados:
                                        titulo = libro[0]
                                        autor = libro[1]
                                        genero = libro[2]
                                        año_publicacion = libro[3]
                                        isbn = libro[4]
                                        fecha_adquisicion = libro[5]
                                        tabla_libros.add_row([titulo, autor, genero, año_publicacion, isbn, fecha_adquisicion])
                                    
                                    # Imprimir la tabla de libros
                                    print("Libros encontrados:")
                                    print(tabla_libros)
                                else:
                                    print("No se encontraron libros con el título especificado.")
                                conn.commit()

                        except Error as e:
                            print("Error:", e)
                        except ValueError:
                            print("Error: Entrada inválida. Por favor, ingrese un valor válido.")
                        except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")


                    if opcion3 == 2:
                        try:
                            with sqlite3.connect("Libreria.db") as conn:
                                cursor = conn.cursor()
                                isbn_buscar = input("Ingrese el ISBN a buscar: ")

                                # Realizar la consulta para obtener los libros con el ISBN especificado y sus detalles con el JOIN
                                cursor.execute("SELECT libros.titulo, autores.nombre_autor, generos.nombre_genero, strftime('%d/%m/%Y', libros.año_publicacion), libros.isbn, strftime('%d/%m/%Y', libros.fecha_adquisicion) FROM libros JOIN autores ON libros.autor_id = autores.id_autor JOIN generos ON libros.genero_id = generos.id_genero WHERE libros.isbn = ?;", (isbn_buscar,))
                                libros_encontrados = cursor.fetchall()

                                if libros_encontrados:
                                    # Crear una instancia de PrettyTable para mostrar los libros
                                    tabla_libros = PrettyTable(["Título", "Autor", "Género", "Año de Publicación", "ISBN", "Fecha de Adquisición"])

                                    # Agregar los registros de libros encontrados a la tabla
                                    for libro in libros_encontrados:
                                        titulo = libro[0]
                                        autor = libro[1]
                                        genero = libro[2]
                                        año_publicacion = libro[3]
                                        isbn = libro[4]
                                        fecha_adquisicion = libro[5]
                                        tabla_libros.add_row([titulo, autor, genero, año_publicacion, isbn, fecha_adquisicion])

                                    # Imprimir la tabla de libros
                                    print("Libros encontrados:")
                                    print(tabla_libros)
                                else:
                                    print("No se encontraron libros con el ISBN especificado.")
                                conn.commit()

                        except Error as e:
                            print("Error:", e)
                        except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")


                    
                    if (opcion3==3):
                        break
                    
                    if (opcion3>3 or opcion3<1):
                        print("Opcion no valida.")
                        continue
                
            if (opcion2==2):
                while True:
                    print("")
                    print("Menu Reportes")
                    print("")
                    print("1.-Catalogo Completo\n2.-Reporte por Autor\n3.-Reporte por genero\n4.-Reporte por Año de Publicacion\n5.-Menu Anterior")
                    opcion4=int(input("Seleccione una opcion: "))
                    print("")
                    if opcion4 == 1:
                        try:
                            with sqlite3.connect("Libreria.db") as conn:
                                cursor = conn.cursor()

                                # Realizar la consulta para obtener todos los libros y sus detalles con el join
                                cursor.execute("SELECT libros.id_libro, libros.titulo, autores.nombre_autor, generos.nombre_genero, strftime('%d/%m/%Y', libros.año_publicacion), libros.isbn, strftime('%d/%m/%Y', libros.fecha_adquisicion) FROM libros JOIN autores ON libros.autor_id = autores.id_autor JOIN generos ON libros.genero_id = generos.id_genero;")
                                libros = cursor.fetchall()

                                # Verificar si existen libros registrados
                                if len(libros) == 0:
                                    print("No hay libros registrados en el catálogo.")
                                else:
                                    # Mostrar el catálogo completo sin encabezado utilizando PrettyTable
                                    tabla_libros = PrettyTable(["ID", "Título", "Autor", "Género", "Año de Publicación", "ISBN", "Fecha de Adquisición"])

                                    for libro in libros:
                                        tabla_libros.add_row(libro)
                                    while True:
                                        print("Catálogo completo de libros:")
                                        print(tabla_libros)
                                        print("¿Quieres exportar los datos?")
                                        print("1.- A Csv")
                                        print("2.- A MsExcel")
                                        print("3.- Salir")
                                        archivo_str = input("Opcion: ")
                                        archivo_generar = int(archivo_str)

                                        if archivo_generar == 1:
                                            with open("libros_catalogo.csv", "w", newline="") as archivo:
                                                grabador = csv.writer(archivo)
                                                grabador.writerow(["ID", "Título", "Autor", "Género", "Año de Publicación", "ISBN", "Fecha de Adquisición"])
                                                grabador.writerows([list(libro) for libro in libros])
                                            print("Los datos se han exportado correctamente a libros_catalogo.csv")
                                            break
                                        elif archivo_generar == 2:
                                            # Creamos un nuevo libro de Excel
                                            wb = openpyxl.Workbook()
                                            hoja_activa = wb.active

                                            # Definimos los encabezados de las columnas
                                            hoja_activa["A1"] = "ID"
                                            hoja_activa["B1"] = "Título"
                                            hoja_activa["C1"] = "Autor"
                                            hoja_activa["D1"] = "Género"
                                            hoja_activa["E1"] = "Fecha de Publicación"
                                            hoja_activa["F1"] = "ISBN"
                                            hoja_activa["G1"] = "Fecha de Adquisición"

                                            # Añadimos los datos del catálogo al archivo de Excel
                                            for row, libro in enumerate(libros, start=2):
                                                hoja_activa.cell(row=row, column=1).value = libro[0]
                                                hoja_activa.cell(row=row, column=2).value = libro[1]
                                                hoja_activa.cell(row=row, column=3).value = libro[2]
                                                hoja_activa.cell(row=row, column=4).value = libro[3]
                                                hoja_activa.cell(row=row, column=5).value = libro[4]
                                                hoja_activa.cell(row=row, column=6).value = libro[5]
                                                hoja_activa.cell(row=row, column=7).value = libro[6]

                                            # Añadimos una fila en blanco
                                            hoja_activa.cell(row=len(libros) + 2, column=1).value = ""
                                            # Añadimos la fecha de actualización
                                            hoja_activa.cell(row=len(libros) + 3, column=1).value = "Última actualización"
                                            hoja_activa.cell(row=len(libros) + 3, column=2).value = datetime.datetime.today().strftime('%d/%m/%Y')

                                            # Guardamos el archivo de Excel
                                            wb.save("reporte_libros_catalogo.xlsx")
                                            print("Los datos se han exportado correctamente a reporte_libros_catalogo.xlsx")
                                            break
                                        elif archivo_generar == 3:
                                            break
                                        else:
                                            continue
                        except Error as e:
                            print("Error:", e)
                        except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                    if opcion4 == 2:
                        try:
                            with sqlite3.connect("Libreria.db") as conn:
                                cursor = conn.cursor()

                                # Obtener la lista de autores disponibles
                                cursor.execute("SELECT * FROM autores;")
                                autores = cursor.fetchall()

                                # Mostrar la lista de autores disponibles
                                print("Autores disponibles:")
                                for autor in autores:
                                    print(f"ID: {autor[0]}, Nombre: {autor[1]}")

                                autor_id = input("Ingrese el ID del autor para generar el reporte: ")

                                # Validar si el ID del autor existe en la base de datos
                                cursor.execute("SELECT * FROM autores WHERE id_autor = ?;", (autor_id,))
                                autor_row = cursor.fetchone()

                                if autor_row is None:
                                    print("El ID del autor no existe. Por favor, ingrese un ID válido.")
                                else:
                                    # Realizar la consulta para obtener los libros del autor seleccionado
                                    cursor.execute("SELECT libros.id_libro, libros.titulo, autores.nombre_autor, generos.nombre_genero, strftime('%d/%m/%Y', libros.año_publicacion), libros.isbn, strftime('%d/%m/%Y', libros.fecha_adquisicion) FROM libros JOIN generos ON libros.genero_id = generos.id_genero JOIN autores ON libros.autor_id = autores.id_autor WHERE libros.autor_id = ?;", (autor_id,))
                                    libros = cursor.fetchall()

                                    # Verificar si existen libros del autor seleccionado
                                    if len(libros) == 0:
                                        print("No hay libros registrados para el autor seleccionado.")
                                    else:
                                        # Mostrar el reporte de libros por autor utilizando PrettyTable
                                        tabla_libros = PrettyTable()
                                        tabla_libros.field_names = ["ID", "Título", "Autor", "Género", "Año de Publicación", "ISBN", "Fecha de Adquisición"]

                                        for libro in libros:
                                            tabla_libros.add_row([libro[0], libro[1], libro[2], libro[3], libro[4], libro[5], libro[6]])

                                        print("Reporte de libros por autor:")
                                        print(tabla_libros)

                                        while True:
                                            print("¿Quieres exportar los datos?")
                                            print("1.- A Csv")
                                            print("2.- A MsExcel")
                                            print("3.- Salir")
                                            archivo_str = input("Opción: ")
                                            archivo_generar = int(archivo_str)

                                            if archivo_generar == 1:
                                                with open("libros_autor.csv", "w", newline="") as archivo:
                                                    grabador = csv.writer(archivo)
                                                    grabador.writerow(["ID", "Título", "Autor", "Género", "Año de Publicación", "ISBN", "Fecha de Adquisición"])
                                                    grabador.writerows([list(libro) for libro in libros])
                                                print("Los datos se han exportado correctamente a libros_autor.csv")
                                                break
                                            elif archivo_generar == 2:
                                                # Creamos un nuevo libro de Excel
                                                wb = openpyxl.Workbook()
                                                hoja_activa = wb.active

                                                # Definimos los encabezados de las columnas
                                                hoja_activa["A1"] = "ID"
                                                hoja_activa["B1"] = "Título"
                                                hoja_activa["C1"] = "Autor"
                                                hoja_activa["D1"] = "Género"
                                                hoja_activa["E1"] = "Año de Publicación"
                                                hoja_activa["F1"] = "ISBN"
                                                hoja_activa["G1"] = "Fecha de Adquisición"

                                                # Añadimos los datos del reporte al archivo de Excel
                                                for row, libro in enumerate(libros, start=2):
                                                    hoja_activa.cell(row=row, column=1).value = libro[0]
                                                    hoja_activa.cell(row=row, column=2).value = libro[1]
                                                    hoja_activa.cell(row=row, column=3).value = libro[2]
                                                    hoja_activa.cell(row=row, column=4).value = libro[3]
                                                    hoja_activa.cell(row=row, column=5).value = libro[4]
                                                    hoja_activa.cell(row=row, column=6).value = libro[5]
                                                    hoja_activa.cell(row=row, column=7).value = libro[6]

                                                # Añadimos una fila en blanco
                                                hoja_activa.cell(row=len(libros) + 2, column=1).value = ""

                                                # Añadimos la fecha de actualización
                                                hoja_activa.cell(row=len(libros) + 3, column=1).value = "Última actualización"
                                                hoja_activa.cell(row=len(libros) + 3, column=2).value = datetime.datetime.today().strftime('%d/%m/%Y')

                                                # Guardamos el archivo de Excel
                                                wb.save("reporte_libros_autor.xlsx")
                                                print("Los datos se han exportado correctamente a reporte_libros_autor.xlsx")
                                                break
                                            elif archivo_generar == 3:
                                                break
                                            else:
                                                continue
                        except Error as e:
                            print("Error:", e)
                        except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                    elif opcion4 == 3:
                        try:
                            with sqlite3.connect("Libreria.db") as conn:
                                cursor = conn.cursor()

                                # Obtener la lista de géneros disponibles
                                cursor.execute("SELECT * FROM generos;")
                                generos = cursor.fetchall()

                                # Mostrar la lista de géneros disponibles
                                print("Géneros disponibles:")
                                for genero in generos:
                                    print(f"ID: {genero[0]}, Nombre: {genero[1]}")

                                genero_id = input("Ingrese el ID del género para generar el reporte: ")

                                # Validar si el ID del género existe en la base de datos
                                cursor.execute("SELECT * FROM generos WHERE id_genero = ?;", (genero_id,))
                                genero_row = cursor.fetchone()

                                if genero_row is None:
                                    print("El ID del género no existe. Por favor, ingrese un ID válido.")
                                else:
                                    # Realizar la consulta para obtener los libros del género seleccionado
                                    cursor.execute("SELECT libros.id_libro, libros.titulo, autores.nombre_autor, generos.nombre_genero, strftime('%d/%m/%Y', libros.año_publicacion), libros.isbn, strftime('%d/%m/%Y', libros.fecha_adquisicion) FROM libros JOIN generos ON libros.genero_id = generos.id_genero JOIN autores ON libros.autor_id = autores.id_autor WHERE libros.genero_id = ?;", (genero_id,))
                                    libros = cursor.fetchall()

                                    # Verificar si existen libros del género seleccionado
                                    if len(libros) == 0:
                                        print("No hay libros registrados para el género seleccionado.")
                                    else:
                                        # Mostrar el reporte de libros por género utilizando PrettyTable
                                        tabla_libros = PrettyTable()
                                        tabla_libros.field_names = ["ID", "Título", "Autor", "Género", "Año de Publicación", "ISBN", "Fecha de Adquisición"]

                                        for libro in libros:
                                            tabla_libros.add_row([libro[0], libro[1], libro[2], libro[3], libro[4], libro[5], libro[6]])

                                        print("Reporte de libros por género:")
                                        print(tabla_libros)

                                        while True:
                                            print("¿Quieres exportar los datos?")
                                            print("1.- A Csv")
                                            print("2.- A MsExcel")
                                            print("3.- Salir")
                                            archivo_str = input("Opción: ")
                                            archivo_generar = int(archivo_str)

                                            if archivo_generar == 1:
                                                with open("libros_genero.csv", "w", newline="") as archivo:
                                                    grabador = csv.writer(archivo)
                                                    grabador.writerow(["ID", "Título", "Autor", "Género", "Año de Publicación", "ISBN", "Fecha de Adquisición"])
                                                    grabador.writerows([list(libro) for libro in libros])
                                                print("Los datos se han exportado correctamente a libros_genero.csv")
                                                break
                                            elif archivo_generar == 2:
                                                # Creamos un nuevo libro de Excel
                                                wb = openpyxl.Workbook()
                                                hoja_activa = wb.active

                                                # Definimos los encabezados de las columnas
                                                hoja_activa["A1"] = "ID"
                                                hoja_activa["B1"] = "Título"
                                                hoja_activa["C1"] = "Autor"
                                                hoja_activa["D1"] = "Género"
                                                hoja_activa["E1"] = "Año de Publicación"
                                                hoja_activa["F1"] = "ISBN"
                                                hoja_activa["G1"] = "Fecha de Adquisición"

                                                # Añadimos los datos del reporte al archivo de Excel
                                                for row, libro in enumerate(libros, start=2):
                                                    hoja_activa.cell(row=row, column=1).value = libro[0]
                                                    hoja_activa.cell(row=row, column=2).value = libro[1]
                                                    hoja_activa.cell(row=row, column=3).value = libro[2]
                                                    hoja_activa.cell(row=row, column=4).value = libro[3]
                                                    hoja_activa.cell(row=row, column=5).value = libro[4]
                                                    hoja_activa.cell(row=row, column=6).value = libro[5]
                                                    hoja_activa.cell(row=row, column=7).value = libro[6]

                                                # Añadimos una fila en blanco
                                                hoja_activa.cell(row=len(libros) + 2, column=1).value = ""

                                                # Añadimos la fecha de actualización
                                                hoja_activa.cell(row=len(libros) + 3, column=1).value = "Última actualización"
                                                hoja_activa.cell(row=len(libros) + 3, column=2).value = datetime.datetime.today().strftime('%d/%m/%Y')

                                                # Guardamos el archivo de Excel
                                                wb.save("reporte_libros_genero.xlsx")
                                                print("Los datos se han exportado correctamente a reporte_libros_genero.xlsx")
                                                break

                                            elif archivo_generar == 3:
                                                break
                                            else:
                                                continue
                        except Error as e:
                            print("Error:", e)
                        except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                    if opcion4 == 4:
                        try:
                            with sqlite3.connect("Libreria.db") as conn:
                                cursor = conn.cursor()
                                año = input("Ingrese el año para generar el reporte: ")

                                # Realizar la consulta para obtener los libros publicados en el año especificado y sus detalles con el JOIN
                                cursor.execute("SELECT libros.titulo, autores.nombre_autor, generos.nombre_genero, strftime('%d/%m/%Y', libros.año_publicacion), libros.isbn, strftime('%d/%m/%Y', libros.fecha_adquisicion) FROM libros JOIN autores ON libros.autor_id = autores.id_autor JOIN generos ON libros.genero_id = generos.id_genero WHERE strftime('%Y', libros.año_publicacion) = ?;", (año,))
                                libros_encontrados = cursor.fetchall()

                                if libros_encontrados:
                                    tabla_libros = PrettyTable(["Título", "Autor", "Género", "Año de Publicación", "ISBN", "Fecha de Adquisición"])

                                    # Agregar los registros de libros encontrados a la tabla
                                    for libro in libros_encontrados:
                                        titulo = libro[0]
                                        autor = libro[1]
                                        genero = libro[2]
                                        año_publicacion = libro[3]
                                        isbn = libro[4]
                                        fecha_adquisicion = libro[5]
                                        tabla_libros.add_row([titulo, autor, genero, año_publicacion, isbn, fecha_adquisicion])
                                    
                                    # Imprimir la tabla de libros
                                    print("Reporte de libros publicados en el año", año)
                                    print(tabla_libros)
                                    
                                    # Exportar los datos a un archivo
                                    while True:
                                        print("¿Quieres exportar los datos?")
                                        print("1.- A CSV")
                                        print("2.- A Excel")
                                        print("3.- Salir")
                                        archivo_str = input("Opción: ")
                                        archivo_generar = int(archivo_str)

                                        if archivo_generar == 1:
                                            with open(f"libros_{año}.csv", "w", newline="") as archivo:
                                                grabador = csv.writer(archivo)
                                                grabador.writerow(["Título", "Autor", "Género", "Año de Publicación", "ISBN", "Fecha de Adquisición"])
                                                grabador.writerows([list(libro) for libro in libros_encontrados])
                                            print(f"Los datos se han exportado correctamente a libros_{año}.csv")
                                            break
                                        elif archivo_generar == 2:
                                            # Creamos un nuevo libro de Excel
                                            wb = openpyxl.Workbook()
                                            hoja_activa = wb.active

                                            # Definimos los encabezados de las columnas
                                            hoja_activa["A1"] = "Título"
                                            hoja_activa["B1"] = "Autor"
                                            hoja_activa["C1"] = "Género"
                                            hoja_activa["D1"] = "Año de Publicación"
                                            hoja_activa["E1"] = "ISBN"
                                            hoja_activa["F1"] = "Fecha de Adquisición"

                                            # Añadimos los datos del reporte al archivo de Excel
                                            for row, libro in enumerate(libros_encontrados, start=2):
                                                hoja_activa.cell(row=row, column=1).value = libro[0]
                                                hoja_activa.cell(row=row, column=2).value = libro[1]
                                                hoja_activa.cell(row=row, column=3).value = libro[2]
                                                hoja_activa.cell(row=row, column=4).value = libro[3]
                                                hoja_activa.cell(row=row, column=5).value = libro[4]
                                                hoja_activa.cell(row=row, column=6).value = libro[5]

                                            # Añadimos una fila en blanco
                                            hoja_activa.cell(row=len(libros_encontrados) + 2, column=1).value = ""

                                            # Añadimos la fecha de actualización
                                            hoja_activa.cell(row=len(libros_encontrados) + 3, column=1).value = "Última actualización"
                                            hoja_activa.cell(row=len(libros_encontrados) + 3, column=2).value = datetime.datetime.today().strftime('%d/%m/%Y')

                                            # Guardamos el archivo de Excel
                                            wb.save(f"reporte_libros_{año}.xlsx")
                                            print(f"Los datos se han exportado correctamente a reporte_libros_{año}.xlsx")
                                            break
                                        elif archivo_generar == 3:
                                            break
                                        else:
                                            continue
                                else:
                                    print("No se encontraron libros publicados en el año especificado.")
                                conn.commit()

                        except Error as e:
                            print("Error:", e)
                        except ValueError:
                            print("Error: Entrada inválida. Por favor, ingrese un valor válido.")
                        except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                    if (opcion4==5):
                        break
                        
                    if (opcion4<1 or opcion4>5):
                        print("Opcion no valida.")
                        continue
               
            if (opcion2==3):
                break   
                          
            if (opcion2<1 or opcion2>3):
                print("Opcion no valida")
                continue
      
    if opcion == 3:
        while True:
            nombre_autor = input("Ingrese el nombre del autor: ")
            if nombre_autor.strip() == "":
                continue

            # Verificar si el autor ya existe en la tabla autores
            try:
                conn = sqlite3.connect("Libreria.db")
                cursor = conn.cursor()
                cursor.execute("SELECT nombre_autor FROM autores WHERE nombre_autor = ?;", (nombre_autor,))
                existing_author = cursor.fetchone()

                if existing_author is not None:
                    print("El autor ya existe. Por favor, ingrese otro nombre.")
                    continue

                # Realizar la inserción en la tabla autores
                cursor.execute("INSERT INTO autores (nombre_autor) VALUES (?);", (nombre_autor,))
                conn.commit()
                print("El autor se ha registrado exitosamente.")
                print(f"La clave asignada fue {cursor.lastrowid}")
                break
            except Error as e:
                print("Error al insertar el autor:", e)
    if opcion == 4:
        while True:
            nombre_genero = input("Ingrese el nombre del género: ")
            if nombre_genero.strip() == "":
                continue

            # Verificar si el género ya existe en la tabla generos
            try:
                conn = sqlite3.connect("Libreria.db")
                cursor = conn.cursor()
                cursor.execute("SELECT nombre_genero FROM generos WHERE nombre_genero = ?;", (nombre_genero,))
                existing_genre = cursor.fetchone()

                if existing_genre is not None:
                    print("El género ya existe. Por favor, ingrese otro nombre.")
                    continue

                # Realizar la inserción en la tabla generos
                cursor.execute("INSERT INTO generos (nombre_genero) VALUES (?);", (nombre_genero,))
                conn.commit()
                print("El género se ha registrado exitosamente.")
                print(f"La clave asignada fue {cursor.lastrowid}")
                break
            except Error as e:
                print("Error al insertar el género:", e)

    if (opcion==5):
        break
     
    if (opcion>5 or opcion<1):
        print("Opcion no valida")
        continue
